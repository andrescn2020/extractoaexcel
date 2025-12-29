import streamlit as st
import io
import pdfplumber
import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule

# Regex para caracteres ilegales en Excel
ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')

def clean_for_excel(text):
    """Elimina caracteres ilegales para Excel y espacios extra"""
    if not text: return ""
    text = str(text)
    text = ILLEGAL_CHARACTERS_RE.sub("", text)
    return text.strip()

def parse_amount(s):
    # Soporta: "1,000.00", "100.00", ".16", "0.43"
    if not s: return 0.0
    try:
        clean = s.replace(",", "")
        if clean.startswith("."): clean = "0" + clean
        return float(clean)
    except: return 0.0

def procesar_hsbc(archivo_pdf):
    """Procesa archivos PDF de HSBC con formato multi-hoja y multi-cuenta"""
    st.info("Procesando archivo HSBC...")

    try:
        # 1. Leer PDF Completo
        with pdfplumber.open(io.BytesIO(archivo_pdf.read())) as pdf:
            texto_completo = ""
            for page in pdf.pages:
                texto_completo += page.extract_text() + "\n"
        
        lineas = texto_completo.splitlines()
        
        # 2. Metadata (Titular, Periodo)
        titular = "Sin Especificar"
        periodo = "Sin Especificar"
        anio_global = "2024"

        # Titular
        for l in lineas[:15]:
            if "SUCURSAL" in l and "(" in l and ")" in l:
                parts = l.split("SUCURSAL")
                titular = parts[0].strip()
                titular = re.sub(r'\(\d+\)', '', titular).strip()
                break
        
        # Periodo
        for l in lineas[:35]:
            if "EXTRACTO DEL" in l:
                match = re.search(r"EXTRACTO DEL\s*(\d{2}/\d{2}/\d{4})\s*AL\s*(\d{2}/\d{2}/\d{4})", l)
                if match:
                    periodo = f"Del {match.group(1)} al {match.group(2)}"
                    anio_global = match.group(1).split("/")[-1]
                    break
        
        # 3. Detectar Cuentas y su Saldo Inicial Global (Resumen al inicio)
        cuentas_info = {} 
        
        start_resumen = -1
        for i, l in enumerate(lineas):
            if "PRODUCTO" in l and "SALDO ANTERIOR" in l:
                start_resumen = i + 1
                break
        
        if start_resumen != -1:
            for l in lineas[start_resumen:]:
                if "DETALLE DE OPERACIONES" in l or "- DETALLE DE OPERACIONES -" in l: break 
                if not l.strip(): continue

                # Regex flexible para capturar numero de cuenta
                match_cuenta = re.search(r"(\d{3,4}-\d-\d{5}-\d)", l)
                
                if match_cuenta:
                    nro_cuenta = match_cuenta.group(1)
                    moneda = "U$S" if "u$s" in l.lower() or "U$S" in l else "$"
                    
                    raw_name = l[:match_cuenta.start()]
                    nombre_prod = re.split(r"(SUCURSAL|SUC|MRNEZ|SLEIL|CENTRO|MICROCENTRO)", raw_name, flags=re.IGNORECASE)[0].strip()
                    if not nombre_prod: nombre_prod = "Cuenta"

                    nums = re.findall(r"(\d{1,3}(?:,\d{3})*\.\d{2})", l)
                    s_ini = 0.0
                    s_fin = 0.0
                    if len(nums) >= 2:
                        s_ini = parse_amount(nums[-2])
                        s_fin = parse_amount(nums[-1])
                    
                    cuentas_info[nro_cuenta] = {
                        "tipo": nombre_prod,
                        "moneda": moneda,
                        "s_ini": s_ini,
                        "s_fin": s_fin,
                        "movimientos": []
                    }

        # 4. Procesar Bloques de Movimientos
        regex_header_cuenta = r"(?:NRO\. )?(\d{3,4}-\d-\d{5}-\d)"
        
        current_account_key = None
        current_saldo_acum = 0.0
        current_date_str = ""
        
        meses = {"ENE":"01","FEB":"02","MAR":"03","ABR":"04","MAY":"05","JUN":"06",
                 "JUL":"07","AGO":"08","SEP":"09","OCT":"10","NOV":"11","DIC":"12"}

        idx = 0
        while idx < len(lineas):
            l = lineas[idx]
            idx += 1
            
            # Detectar Cabecera de Cuenta
            match_head = re.search(regex_header_cuenta, l)
            if match_head and ("CUENTA" in l or "CAJA" in l or "WPB" in l): 
                nro = match_head.group(1)
                if nro in cuentas_info:
                    current_account_key = nro
                    current_saldo_acum = cuentas_info[nro]["s_ini"]
                    current_date_str = "" 
                continue
            
            if current_account_key:
                if "DETALLE DE TITULARIDAD" in l or "CALCULO DE INTERESES" in l:
                    current_account_key = None
                    continue
                if "- SALDO ANTERIOR" in l or "- SALDO FINAL" in l: continue
                if "FECHA" in l and "SALDO" in l: continue
                if "HOJA" in l and "DE" in l: continue
                
                # --- PARSEO MOVIMIENTO ---
                
                # 1. Fecha?
                match_fecha = re.match(r"(\d{2}-[A-Z]{3})", l.strip())
                if match_fecha:
                    fecha_raw = match_fecha.group(1)
                    dia, mes_txt = fecha_raw.split("-")
                    current_date_str = f"{dia}/{meses.get(mes_txt,'01')}/{anio_global}"
                
                # 2. Buscar numeros decimales al final
                nums = re.findall(r"((?:\d{1,3}(?:,\d{3})*)?\.\d{2})", l)
                nums = [n for n in nums if n.strip() and n != "."]
                
                if len(nums) >= 2: 
                    saldo_str = nums[-1]
                    importe_str = nums[-2]
                    
                    saldo_linea = parse_amount(saldo_str)
                    importe_visual = parse_amount(importe_str)
                    
                    diff = saldo_linea - current_saldo_acum
                    
                    match_diff = abs(abs(diff) - importe_visual) < 1.0
                    
                    if match_diff:
                        importe_real = diff
                        current_saldo_acum = saldo_linea
                        
                        try:
                            idx_imp = l.rfind(importe_str)
                            if idx_imp != -1:
                                desc_raw = l[:idx_imp]
                            else:
                                desc_raw = "Movimiento"
                        except: desc_raw = "Movimiento"
                        
                        if match_fecha:
                            desc_raw = desc_raw[match_fecha.end():]
                            
                        desc = re.sub(r"^\s*-\s*", "", desc_raw).strip()
                        desc = re.sub(r"\d+$", "", desc).strip()
                        
                        if not desc: desc = "Movimiento"

                        cuentas_info[current_account_key]["movimientos"].append({
                            "Fecha": current_date_str or periodo,
                            "Descripcion": desc,
                            "Importe": importe_real
                        })
                        continue 
                
                # Descripcion extra?
                clean_l = l.strip()
                if clean_l and len(clean_l) > 3:
                     if re.search(r"\.\d{2}$", clean_l): continue 
                     
                     movs = cuentas_info[current_account_key]["movimientos"]
                     if movs:
                        txt_add = re.sub(r"^\s*-\s*", "", clean_l).strip()
                        if txt_add and txt_add not in movs[-1]["Descripcion"]:
                            movs[-1]["Descripcion"] += " " + txt_add
        
    except Exception as e:
        import traceback
        st.error(f"Error procesando HSBC: {e}")
        print(traceback.format_exc())
        return None

    return generar_excel_hsbc(cuentas_info, titular, periodo, anio_global)

def generar_excel_hsbc(cuentas_info, titular, periodo, anio_global):
    output = io.BytesIO()
    wb = Workbook()
    if "Sheet" in wb.sheetnames: del wb["Sheet"]
    
    # Estilos
    color_bg_main = "DB0011"
    color_txt_main = "FFFFFF"
    thin = Border(left=Side(style='thin',color="A6A6A6"), right=Side(style='thin',color="A6A6A6"), top=Side(style='thin',color="A6A6A6"), bottom=Side(style='thin',color="A6A6A6"))
    head_c = PatternFill(start_color="00B050", end_color="00B050", fill_type="solid")
    head_d = PatternFill(start_color="C00000", end_color="C00000", fill_type="solid")
    fill_c = PatternFill(start_color="EBF1DE", end_color="EBF1DE", fill_type="solid")
    fill_d = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")
    row_c = PatternFill(start_color="F2F9F1", end_color="F2F9F1", fill_type="solid")
    row_d = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")
    red_f = Font(color='9C0006', bold=True)
    red_bg = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')

    for nro, data in cuentas_info.items():
        if not data["movimientos"] and data["s_ini"] == data["s_fin"]: continue
        
        name = f"{data['tipo'][:10]} {data['moneda']}".replace("/","").strip()
        count = 1
        orig = name
        while name in wb.sheetnames: name = f"{orig} {count}"; count+=1
        
        ws = wb.create_sheet(title=name)
        ws.sheet_view.showGridLines = False
        fmt = '"$ "#,##0.00' if data["moneda"]=="$" else '"U$S "#,##0.00'
        
        df = pd.DataFrame(data["movimientos"])
        if df.empty:
            cred, deb = pd.DataFrame(), pd.DataFrame()
        else:
            cred = df[df["Importe"] > 0].copy()
            deb = df[df["Importe"] < 0].copy()
            deb["Importe"] = deb["Importe"].abs()

        ws.merge_cells("A1:G1"); ws["A1"]=f"REPORTE HSBC - {data['tipo']} - {titular}"; ws["A1"].font=Font(size=14,bold=True,color="FFFFFF"); ws["A1"].fill=PatternFill(start_color=color_bg_main,end_color=color_bg_main,fill_type="solid"); ws["A1"].alignment=Alignment(horizontal="center")
        
        ws["A3"]="SALDO INICIAL"; ws["B3"]=data["s_ini"]; ws["B3"].number_format=fmt
        ws["A4"]="SALDO FINAL"; ws["B4"]=data["s_fin"]; ws["B4"].number_format=fmt
        ws["D3"]="TITULAR"; ws["E3"]=titular
        ws["D4"]="PERIODO"; ws["E4"]=periodo
        
        ws["D6"]="CONTROL"; ws["D7"]=0; ws["D7"].font=Font(bold=True)
        ws.conditional_formatting.add('D7', CellIsRule(operator='notEqual', formula=['0'], stopIfTrue=True, fill=red_bg, font=red_f))
        
        r = 10
        ws.merge_cells(f"A{r}:C{r}"); ws[f"A{r}"]="CRÉDITOS"; ws[f"A{r}"].fill=head_c; ws[f"A{r}"].font=Font(color="FFFFFF",bold=True)
        ws.merge_cells(f"E{r}:G{r}"); ws[f"E{r}"]="DÉBITOS"; ws[f"E{r}"].fill=head_d; ws[f"E{r}"].font=Font(color="FFFFFF",bold=True)
        
        r+=1
        for c, t in zip(["A","B","C","E","F","G"], ["Fecha","Desc","Imp","Fecha","Desc","Imp"]):
            ws[f"{c}{r}"]=t; ws[f"{c}{r}"].border=thin; ws[f"{c}{r}"].font=Font(bold=True)
            if c in "ABC": ws[f"{c}{r}"].fill=fill_c
            else: ws[f"{c}{r}"].fill=fill_d
            
        r+=1
        s_c = r
        if not cred.empty:
            for _, row in cred.iterrows():
                ws[f"A{r}"]=row["Fecha"]; ws[f"B{r}"]=row["Descripcion"]; ws[f"C{r}"]=row["Importe"]
                ws[f"C{r}"].number_format=fmt; 
                for x in "ABC": ws[f"{x}{r}"].fill=row_c; ws[f"{x}{r}"].border=thin
                r+=1
        t_c = r; ws[f"C{t_c}"]=f"=SUM(C{s_c}:C{t_c-1})"; ws[f"C{t_c}"].number_format=fmt
        
        r_d = 12
        s_d = r_d
        if not deb.empty:
            for _, row in deb.iterrows():
                ws[f"E{r_d}"]=row["Fecha"]; ws[f"F{r_d}"]=row["Descripcion"]; ws[f"G{r_d}"]=row["Importe"]
                ws[f"G{r_d}"].number_format=fmt
                for x in "EFG": ws[f"{x}{r_d}"].fill=row_d; ws[f"{x}{r_d}"].border=thin
                r_d+=1
        t_d = r_d; ws[f"G{t_d}"]=f"=SUM(G{s_d}:G{t_d-1})"; ws[f"G{t_d}"].number_format=fmt
        
        ws["D7"] = f"=ROUND(B3+C{t_c}-G{t_d}-B4, 2)"
        
        ws.column_dimensions["B"].width=40; ws.column_dimensions["F"].width=40
        ws.column_dimensions["C"].width=18; ws.column_dimensions["G"].width=18
        
    wb.save(output)
    output.seek(0)
    return output.getvalue()
